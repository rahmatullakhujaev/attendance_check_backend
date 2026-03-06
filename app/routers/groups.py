import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from fastapi.responses import StreamingResponse
import io
from datetime import timezone
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_instructor
from app.models import Group, Participant, Student, Lecture, Participation
from app.schemas import GroupCreate, GroupUpdate, GroupResponse

router = APIRouter(prefix="/groups", tags=["groups"])


@router.get("/", response_model=list[GroupResponse])
def list_groups(
    db: Session = Depends(get_db),
    instructor=Depends(get_current_instructor),
):
    """List all groups belonging to the current instructor."""
    return db.query(Group).filter(Group.instructor_id == instructor.id).all()


@router.post("/", response_model=GroupResponse, status_code=201)
def create_group(
    payload: GroupCreate,
    db: Session = Depends(get_db),
    instructor=Depends(get_current_instructor),
):
    group = Group(
        instructor_id=instructor.id,
        name=payload.name,
        course=payload.course,
    )
    db.add(group)
    db.commit()
    db.refresh(group)
    return group


@router.get("/{group_id}", response_model=GroupResponse)
def get_group(
    group_id: int,
    db: Session = Depends(get_db),
    instructor=Depends(get_current_instructor),
):
    group = _get_group_or_404(group_id, instructor.id, db)
    return group


@router.patch("/{group_id}", response_model=GroupResponse)
def update_group(
    group_id: int,
    payload: GroupUpdate,
    db: Session = Depends(get_db),
    instructor=Depends(get_current_instructor),
):
    group = _get_group_or_404(group_id, instructor.id, db)
    if payload.name is not None:
        group.name = payload.name
    if payload.course is not None:
        group.course = payload.course
    db.commit()
    db.refresh(group)
    return group


@router.delete("/{group_id}", status_code=204)
def delete_group(
    group_id: int,
    db: Session = Depends(get_db),
    instructor=Depends(get_current_instructor),
):
    group = _get_group_or_404(group_id, instructor.id, db)
    db.delete(group)
    db.commit()


@router.get("/{group_id}/participants", response_model=list)
def list_participants(
    group_id: int,
    db: Session = Depends(get_db),
    instructor=Depends(get_current_instructor),
):
    """List all students in a group."""
    _get_group_or_404(group_id, instructor.id, db)
    participants = (
        db.query(Participant)
        .filter(Participant.group_id == group_id)
        .all()
    )
    return [
        {
            "id": p.student.id,  # internal DB id for photo upload
            "student_id": p.student.student_id,  # university ID
            "first_name": p.student.first_name,
            "last_name": p.student.last_name,
            "photo": p.student.photo,
            "has_embedding": p.student.embedding is not None,
        }
        for p in participants
    ]

@router.post("/{group_id}/participants/add", status_code=201)
def add_participant(
    group_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    instructor=Depends(get_current_instructor),
):
    group = _get_group_or_404(group_id, instructor.id, db)
    student = db.query(Student).filter(Student.id == payload["student_id"]).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    existing = db.query(Participant).filter(
        Participant.group_id == group_id,
        Participant.student_id == student.id,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Student already in group")
    db.add(Participant(group_id=group_id, student_id=student.id))
    db.commit()
    return {"ok": True}

@router.get("/{group_id}/statistics/excel")
def export_statistics_excel(
    group_id: int,
    db: Session = Depends(get_db),
    instructor=Depends(get_current_instructor),
):
    group = _get_group_or_404(group_id, instructor.id, db)

    # Get all lectures for this group ordered by time
    lectures = db.query(Lecture).filter(
        Lecture.group_id == group_id
    ).order_by(Lecture.time).all()

    # Get all participants
    participants = db.query(Participant).filter(
        Participant.group_id == group_id
    ).all()

    # Build attendance matrix
    # {participant_id: {lecture_id: status}}
    matrix = {}
    for p in participants:
        matrix[p.id] = {}

    for lecture in lectures:
        participations = db.query(Participation).filter(
            Participation.lecture_id == lecture.id
        ).all()
        for part in participations:
            if part.participant_id in matrix:
                matrix[part.participant_id][lecture.id] = part.status.value

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Styles
    header_fill    = PatternFill("solid", fgColor="6C63FF")
    present_fill   = PatternFill("solid", fgColor="1E7E34")
    absent_fill    = PatternFill("solid", fgColor="C0392B")
    late_fill      = PatternFill("solid", fgColor="E67E22")
    header_font    = Font(bold=True, color="FFFFFF")
    center         = Alignment(horizontal="center", vertical="center")

    # Header row
    headers = ["#", "Full Name", "Student ID"]
    for lecture in lectures:
        t = lecture.time
        headers.append(
            f"{t.day:02d}/{t.month:02d}/{t.year}\n{t.hour:02d}:{t.minute:02d}"
        )
    headers.append("Present")
    headers.append("Absent")
    headers.append("Late")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center

    ws.row_dimensions[1].height = 40

    # Data rows
    for row_idx, participant in enumerate(participants, 2):
        student = participant.student
        attendance = matrix.get(participant.id, {})

        present_count = sum(1 for s in attendance.values() if s == "present")
        absent_count  = sum(1 for s in attendance.values() if s == "absent")
        late_count    = sum(1 for s in attendance.values() if s == "late")

        # # column
        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center

        # Full name
        ws.cell(row=row_idx, column=2,
                value=f"{student.first_name} {student.last_name}")

        # Student ID
        ws.cell(row=row_idx, column=3,
                value=student.student_id).alignment = center

        # Lecture columns
        for col_idx, lecture in enumerate(lectures, 4):
            status = attendance.get(lecture.id, "absent")
            if status == "present":
                symbol = "+"
                fill   = present_fill
            elif status == "late":
                symbol = "late"
                fill   = late_fill
            else:
                symbol = "-"
                fill   = absent_fill

            cell           = ws.cell(row=row_idx, column=col_idx, value=symbol)
            cell.fill      = fill
            cell.font      = Font(color="FFFFFF", bold=True)
            cell.alignment = center

        # Summary columns
        ws.cell(row=row_idx, column=len(lectures) + 4,
                value=present_count).alignment = center
        ws.cell(row=row_idx, column=len(lectures) + 5,
                value=absent_count).alignment  = center
        ws.cell(row=row_idx, column=len(lectures) + 6,
                value=late_count).alignment    = center

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    for col_idx in range(4, len(lectures) + 4):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = 14
    ws.column_dimensions[
        openpyxl.utils.get_column_letter(len(lectures) + 4)
    ].width = 10
    ws.column_dimensions[
        openpyxl.utils.get_column_letter(len(lectures) + 5)
    ].width = 10
    ws.column_dimensions[
        openpyxl.utils.get_column_letter(len(lectures) + 6)
    ].width = 10

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{group.name}_attendance.xlsx".replace(" ", "_")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _get_group_or_404(group_id: int, instructor_id: int, db: Session) -> Group:
    group = db.query(Group).filter(
        Group.id == group_id,
        Group.instructor_id == instructor_id,
    ).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    return group